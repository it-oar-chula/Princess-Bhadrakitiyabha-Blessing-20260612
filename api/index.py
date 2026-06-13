from __future__ import annotations

import csv
import io
import json
import os
import secrets
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from textwrap import dedent
from typing import Optional

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

ROOT = Path(__file__).resolve().parent.parent
DB_PATH_RAW = os.getenv("DB_PATH", str(ROOT / "data" / "signatures.db"))
DB_PATH = Path(DB_PATH_RAW)
if not DB_PATH.is_absolute():
    DB_PATH = (ROOT / DB_PATH).resolve()
else:
    DB_PATH = DB_PATH.resolve()

ADMIN_CODE = os.getenv("ADMIN_CODE", "change-me")
APP_TITLE = os.getenv("APP_TITLE", "ร่วมลงนามถวายความอาลัย")
EXPORT_BASENAME = os.getenv("EXPORT_BASENAME", "20260612-Bhadrakitiyabha-Blessing")
BANGKOK_TZ = timezone(timedelta(hours=7))

PHRASES = [
    "เสด็จสู่แดนสรวง\nไทยทั้งปวงน้อมสำนึกในพระกรุณาธิคุณตราบนิรันดร์",
    "สถิตในใจปวงประชาตราบนิรันดร์\nด้วยสำนึกในพระกรุณาธิคุณเป็นล้นพ้น",
    "สถิตในใจไทยนิรันดร์\nน้อมสำนึกในพระกรุณาธิคุณเป็นล้นพ้น",
    "ผองไทยน้อมสำนึกในพระกรุณาธิคุณตราบนิจนิรันดร์",
    "ผองพสกนิกรน้อมสำนึกในพระกรุณาธิคุณตราบนิจนิรันดร์",
    "น้อมสำนึกในพระกรุณาธิคุณตราบนิรันดร์",
    "พระกรุณาธิคุณจารึกในใจไทยตราบนิจนิรันดร์",
    "สถิตกลางใจปวงประชา\nด้วยสำนึกในพระกรุณาธิคุณตราบนิรันดร์",
]

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS signatures (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at_ms INTEGER NOT NULL,
    submitted_at TEXT NOT NULL,
    name TEXT NOT NULL,
    phrase_index INTEGER NOT NULL,
    phrase TEXT NOT NULL
)
"""

# ย้ายข้อมูลจาก schema เดิม (id แบบ UUID + คอลัมน์ status/note) มาเป็น id เลขรัน 1,2,3...
MIGRATE_SQL = """
CREATE TABLE signatures_new (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at_ms INTEGER NOT NULL,
    submitted_at TEXT NOT NULL,
    name TEXT NOT NULL,
    phrase_index INTEGER NOT NULL,
    phrase TEXT NOT NULL
);
INSERT INTO signatures_new (created_at_ms, submitted_at, name, phrase_index, phrase)
    SELECT created_at_ms, submitted_at, name, phrase_index, phrase
    FROM signatures
    ORDER BY created_at_ms ASC, rowid ASC;
DROP TABLE signatures;
ALTER TABLE signatures_new RENAME TO signatures;
"""

app = FastAPI(title=APP_TITLE)


class SignatureCreate(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    phrase_index: Optional[int] = Field(default=None, ge=0, le=len(PHRASES) - 1)
    phrase: Optional[str] = Field(default=None, max_length=300)


class SignatureUpdate(BaseModel):
    name: Optional[str] = Field(default=None, max_length=120)
    phrase_index: Optional[int] = Field(default=None, ge=0, le=len(PHRASES) - 1)


class BulkDelete(BaseModel):
    ids: list[int] = Field(default_factory=list)


def now_bangkok() -> datetime:
    return datetime.now(BANGKOK_TZ)


def normalize_text(value: str) -> str:
    return " ".join(str(value).split()).strip()


def ensure_database() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH, timeout=10) as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA foreign_keys=ON;")
        conn.execute(CREATE_TABLE_SQL)

        # ถ้าเป็น schema เดิม (มีคอลัมน์ status/note) ให้ย้ายข้อมูลมา schema ใหม่ + ออกเลขรันใหม่
        columns = [row[1] for row in conn.execute("PRAGMA table_info(signatures)").fetchall()]
        if "status" in columns or "note" in columns:
            conn.executescript(MIGRATE_SQL)

        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_signatures_created_at_ms "
            "ON signatures(created_at_ms DESC)"
        )
        conn.commit()


def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    conn.execute("PRAGMA busy_timeout=5000;")
    return conn


def row_to_entry(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "at": row["created_at_ms"],
        "submitted_at": row["submitted_at"],
        "name": row["name"],
        "phrase_index": row["phrase_index"],
        "phrase": row["phrase"],
    }


def resolve_phrase(payload: SignatureCreate) -> tuple[int, str]:
    if payload.phrase_index is not None:
        return payload.phrase_index, PHRASES[payload.phrase_index]

    if payload.phrase:
        normalized = payload.phrase.replace("\r\n", "\n").strip()
        if normalized in PHRASES:
            return PHRASES.index(normalized), normalized

    raise HTTPException(status_code=400, detail="phrase_index is required")


def fetch_entries(order: str = "DESC") -> list[dict]:
    if order not in {"ASC", "DESC"}:
        raise ValueError("order must be ASC or DESC")

    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT id, created_at_ms, submitted_at, name, phrase_index, phrase
            FROM signatures
            ORDER BY created_at_ms {order}, id {order}
            """
        ).fetchall()
    return [row_to_entry(row) for row in rows]


def fetch_summary() -> dict:
    with get_connection() as conn:
        total = conn.execute("SELECT COUNT(*) AS count FROM signatures").fetchone()["count"]
        phrase_rows = conn.execute(
            """
            SELECT phrase_index, COUNT(*) AS count
            FROM signatures
            GROUP BY phrase_index
            """
        ).fetchall()

    counts = {row["phrase_index"]: row["count"] for row in phrase_rows}
    phrases = [
        {
            "phrase_index": index,
            "phrase": phrase,
            "count": int(counts.get(index, 0)),
        }
        for index, phrase in enumerate(PHRASES)
    ]
    return {"total": int(total), "phrases": phrases}


def summary_from_entries(entries: list[dict]) -> dict:
    counts = {index: 0 for index in range(len(PHRASES))}
    for entry in entries:
        phrase_index = entry.get("phrase_index")
        if isinstance(phrase_index, int) and phrase_index in counts:
            counts[phrase_index] += 1

    phrases = [
        {
            "phrase_index": index,
            "phrase": phrase,
            "count": counts[index],
        }
        for index, phrase in enumerate(PHRASES)
    ]
    return {"total": len(entries), "phrases": phrases}


def require_admin(code: Optional[str]) -> None:
    if not code or not secrets.compare_digest(code, ADMIN_CODE):
        raise HTTPException(status_code=401, detail="Invalid admin code")


EXPORT_HEADERS = ["id", "submitted_at", "name", "phrase_index", "phrase"]


def entry_to_row(entry: dict) -> list[object]:
    return [
        entry["id"],
        entry["submitted_at"],
        entry["name"],
        entry["phrase_index"],
        entry["phrase"],
    ]


def build_csv_bytes(entries: list[dict]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_HEADERS)
    for entry in entries:
        writer.writerow(entry_to_row(entry))
    return buffer.getvalue().encode("utf-8-sig")


def set_sheet_widths(worksheet, rows: list[list[object]], headers: list[str]) -> None:
    widths = [len(str(header)) for header in headers]
    for row in rows:
        for index, value in enumerate(row):
            widths[index] = max(widths[index], len(str(value)))

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(width + 4, 48)


def build_xlsx_bytes(entries: list[dict]) -> bytes:
    summary = summary_from_entries(entries)
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "entries"
    headers = EXPORT_HEADERS
    sheet.append(headers)
    entry_rows: list[list[object]] = []
    for entry in entries:
        row = entry_to_row(entry)
        entry_rows.append(row)
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_sheet_widths(sheet, entry_rows, headers)

    summary_sheet = workbook.create_sheet("summary")
    summary_sheet.append(["total_entries", summary["total"]])
    summary_sheet.append([])
    summary_sheet.append(["phrase_index", "phrase", "count"])
    summary_rows: list[list[object]] = []
    for item in summary["phrases"]:
        row = [item["phrase_index"], item["phrase"], item["count"]]
        summary_rows.append(row)
        summary_sheet.append(row)
    summary_sheet.freeze_panes = "A3"
    set_sheet_widths(summary_sheet, summary_rows, ["phrase_index", "phrase", "count"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@app.on_event("startup")
def startup() -> None:
    ensure_database()


@app.get("/api/health")
def health() -> dict:
    return {
        "status": "online",
        "records": fetch_summary()["total"],
        "database": str(DB_PATH.name),
    }


@app.get("/api/signatures")
def list_signatures() -> list[dict]:
    return fetch_entries(order="DESC")


def insert_signature(name: str, phrase_index: int, phrase: str) -> dict:
    created_at = now_bangkok()
    created_at_ms = int(created_at.timestamp() * 1000)
    submitted_at = created_at.isoformat(timespec="seconds")

    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO signatures (
                created_at_ms, submitted_at, name, phrase_index, phrase
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (created_at_ms, submitted_at, name, phrase_index, phrase),
        )
        new_id = cursor.lastrowid
        conn.commit()

    return {
        "id": new_id,
        "at": created_at_ms,
        "submitted_at": submitted_at,
        "name": name,
        "phrase_index": phrase_index,
        "phrase": phrase,
    }


@app.post("/api/signatures")
def create_signature(payload: SignatureCreate) -> dict:
    name = normalize_text(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    phrase_index, phrase = resolve_phrase(payload)
    return {"ok": True, "entry": insert_signature(name, phrase_index, phrase)}


@app.get("/api/admin/summary")
def admin_summary(code: str = Query(default="")) -> dict:
    require_admin(code)
    return fetch_summary()


@app.get("/api/admin/export.csv")
def admin_export_csv(code: str = Query(default="")) -> Response:
    require_admin(code)
    entries = fetch_entries(order="ASC")
    payload = build_csv_bytes(entries)
    filename = f"{EXPORT_BASENAME}.csv"
    return Response(
        content=payload,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/export.xlsx")
def admin_export_xlsx(code: str = Query(default="")) -> Response:
    require_admin(code)
    entries = fetch_entries(order="ASC")
    payload = build_xlsx_bytes(entries)
    filename = f"{EXPORT_BASENAME}.xlsx"
    return Response(
        content=payload,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/admin/entries")
def admin_entries(code: str = Query(default="")) -> list[dict]:
    require_admin(code)
    return fetch_entries(order="ASC")


@app.post("/api/admin/signatures")
def admin_add_signature(payload: SignatureCreate, code: str = Query(default="")) -> dict:
    require_admin(code)
    name = normalize_text(payload.name)
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    phrase_index, phrase = resolve_phrase(payload)
    return {"ok": True, "entry": insert_signature(name, phrase_index, phrase)}


@app.patch("/api/admin/signatures/{entry_id}")
def admin_update_signature(
    entry_id: int, payload: SignatureUpdate, code: str = Query(default="")
) -> dict:
    require_admin(code)

    fields: list[str] = []
    params: list[object] = []
    if payload.name is not None:
        name = normalize_text(payload.name)
        if not name:
            raise HTTPException(status_code=400, detail="name cannot be empty")
        fields.append("name = ?")
        params.append(name)
    if payload.phrase_index is not None:
        fields.append("phrase_index = ?")
        params.append(payload.phrase_index)
        fields.append("phrase = ?")
        params.append(PHRASES[payload.phrase_index])

    if not fields:
        raise HTTPException(status_code=400, detail="nothing to update")

    params.append(entry_id)
    with get_connection() as conn:
        cursor = conn.execute(
            f"UPDATE signatures SET {', '.join(fields)} WHERE id = ?", params
        )
        conn.commit()
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="entry not found")

    return {"ok": True}


@app.delete("/api/admin/signatures/{entry_id}")
def admin_delete_signature(entry_id: int, code: str = Query(default="")) -> dict:
    require_admin(code)
    with get_connection() as conn:
        cursor = conn.execute("DELETE FROM signatures WHERE id = ?", (entry_id,))
        conn.commit()
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="entry not found")
    return {"ok": True}


@app.get("/api/admin/verify")
def admin_verify(code: str = Query(default="")) -> dict:
    require_admin(code)
    return {"ok": True}


@app.post("/api/admin/signatures/bulk-delete")
def admin_bulk_delete(payload: BulkDelete, code: str = Query(default="")) -> dict:
    require_admin(code)
    ids = [int(i) for i in payload.ids]
    if not ids:
        raise HTTPException(status_code=400, detail="no ids provided")
    placeholders = ",".join("?" for _ in ids)
    with get_connection() as conn:
        cursor = conn.execute(
            f"DELETE FROM signatures WHERE id IN ({placeholders})", ids
        )
        conn.commit()
    return {"ok": True, "deleted": cursor.rowcount}


@app.get("/admin", response_class=HTMLResponse)
def admin_page() -> str:
    phrases_json = json.dumps([p.replace("\n", " ") for p in PHRASES], ensure_ascii=False)
    html = dedent(
        f"""
        <!doctype html>
        <html lang="th">
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <title>{APP_TITLE} - Admin</title>
          <style>
            :root {{
              color-scheme: dark;
              --bg: #070708;
              --panel: #121214;
              --line: #303036;
              --ink: #f3f0ea;
              --muted: #a39d93;
              --accent: #d5cdbd;
            }}
            * {{ box-sizing: border-box; }}
            body {{
              margin: 0;
              font-family: system-ui, -apple-system, Segoe UI, sans-serif;
              background:
                radial-gradient(900px 500px at 50% 0%, rgba(255,255,255,0.08), transparent 62%),
                linear-gradient(180deg, #050506 0%, #0f0f11 100%);
              color: var(--ink);
            }}
            main {{
              max-width: 1200px;
              margin: 0 auto;
              padding: 32px 24px 56px;
            }}
            h1 {{
              margin: 0 0 8px;
              font-size: 1.7rem;
              font-weight: 650;
            }}
            p {{ color: var(--muted); line-height: 1.6; }}
            .panel {{
              margin-top: 24px;
              border: 1px solid var(--line);
              background: rgba(18, 18, 20, 0.96);
              border-radius: 8px;
              padding: 20px;
            }}
            label {{
              display: block;
              margin-bottom: 8px;
              color: var(--accent);
              font-size: 0.95rem;
            }}
            input {{
              width: 100%;
              padding: 12px 14px;
              border: 1px solid var(--line);
              background: #09090a;
              color: var(--ink);
              border-radius: 6px;
              font: inherit;
            }}
            .row {{
              display: flex;
              gap: 10px;
              flex-wrap: wrap;
              margin-top: 14px;
            }}
            button {{
              padding: 12px 16px;
              border: 1px solid var(--accent);
              border-radius: 6px;
              background: var(--accent);
              color: #111;
              font: inherit;
              cursor: pointer;
            }}
            button.secondary {{
              background: transparent;
              color: var(--ink);
              border-color: var(--line);
            }}
            pre {{
              margin-top: 16px;
              padding: 16px;
              white-space: pre-wrap;
              border: 1px solid var(--line);
              background: #0b0b0c;
              border-radius: 6px;
              min-height: 88px;
            }}
            small {{
              display: block;
              margin-top: 12px;
              color: var(--muted);
            }}
            h2 {{ margin: 0 0 6px; font-size: 1.2rem; font-weight: 600; }}
            input[type="checkbox"] {{ width: auto; padding: 0; cursor: pointer; }}
            select {{
              padding: 10px 12px;
              border: 1px solid var(--line);
              background: #09090a;
              color: var(--ink);
              border-radius: 6px;
              font: inherit;
              max-width: 100%;
            }}
            .login-card {{ max-width: 460px; margin: 48px auto 0; }}
            #adminContent[hidden] {{ display: none; }}
            .topbar {{ display: flex; justify-content: space-between; align-items: center; gap: 12px; margin-top: 6px; }}
            .toolbar {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; margin-top: 16px; }}
            .toolbar .search {{ flex: 1; min-width: 260px; }}
            .addbox {{
              display: grid;
              grid-template-columns: 1fr 1.3fr auto;
              gap: 10px;
              margin-top: 14px;
            }}
            .addbox > * {{ min-width: 0; }}
            .addbox input, .addbox select {{ width: 100%; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 16px; font-size: 0.92rem; table-layout: fixed; }}
            th, td {{ padding: 8px 8px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: middle; }}
            th {{ color: var(--accent); font-weight: 600; }}
            th.cb, td.cb {{ width: 36px; text-align: center; }}
            th:nth-child(2), td:nth-child(2) {{ width: 46px; }}
            th:nth-child(3), td:nth-child(3) {{ width: 28%; }}
            th:nth-child(5), td:nth-child(5) {{ width: 150px; color: var(--muted); font-size: 0.82rem; }}
            th:nth-child(6), td:nth-child(6) {{ width: 150px; }}
            td input.cell-name {{ width: 100%; padding: 8px 10px; }}
            td select {{ width: 100%; }}
            td.actions {{ white-space: nowrap; }}
            td.actions button {{ padding: 7px 12px; font-size: 0.85rem; }}
            button.danger {{ background: #7c2d2d; border-color: #7c2d2d; color: #fbeaea; }}
            .muted-id {{ color: var(--muted); }}
          </style>
        </head>
        <body>
          <main>
            <h1>{APP_TITLE}</h1>

            <section class="panel login-card" id="loginCard">
              <label for="loginCode">Admin password</label>
              <input id="loginCode" type="password" autocomplete="current-password" placeholder="ใส่รหัสผ่านผู้ดูแล">
              <div class="row">
                <button type="button" id="loginBtn">เข้าสู่ระบบ</button>
              </div>
              <small id="loginError" style="color:#e0a0a0; min-height:1.2em;"></small>
            </section>

            <div id="adminContent" hidden>
              <div class="topbar">
                <p>หน้าผู้ดูแล — โหลด CSV/XLSX, ดูสรุป และจัดการรายชื่อ</p>
                <button type="button" class="secondary" id="logoutBtn">ออกจากระบบ</button>
              </div>

              <section class="panel">
                <div class="row">
                  <button type="button" id="downloadCsv">Download CSV</button>
                  <button type="button" id="downloadXlsx">Download XLSX</button>
                  <button type="button" class="secondary" id="loadSummary">Load Summary</button>
                </div>
                <pre id="status">พร้อมใช้งาน</pre>
              </section>

              <section class="panel">
                <h2>จัดการรายชื่อ (เพิ่ม / แก้ไข / ลบ)</h2>

                <div class="addbox">
                  <input id="newName" type="text" maxlength="120" placeholder="ชื่อผู้ลงนาม">
                  <select id="newPhrase"></select>
                  <button type="button" id="addBtn">+ เพิ่มรายชื่อ</button>
                </div>

                <div class="toolbar">
                  <button type="button" class="secondary" id="loadEntries">โหลดรายชื่อทั้งหมด</button>
                  <input class="search" id="searchInput" type="text" placeholder="ค้นหาจากชื่อ หรือ วันที่...">
                  <button type="button" class="danger" id="bulkDeleteBtn">ลบที่เลือก (0)</button>
                  <span id="mgrStatus" style="color:var(--muted);"></span>
                </div>

                <div id="tableWrap"></div>
              </section>
            </div>
          </main>

          <script>
            const statusEl = document.getElementById('status');
            const storageKey = 'blessing-admin-code';
            let adminCode = '';
            let allEntries = [];

            const PHRASES = {phrases_json};

            function getCode() {{ return adminCode; }}
            function setMgr(m) {{ document.getElementById('mgrStatus').textContent = m; }}
            function fmtTime(ms) {{
              try {{ return new Date(ms).toLocaleString('th-TH'); }} catch (e) {{ return ''; }}
            }}

            // ---------- login gate ----------
            async function doLogin(silent) {{
              const errEl = document.getElementById('loginError');
              const code = document.getElementById('loginCode').value.trim();
              errEl.textContent = '';
              if (!code) {{ if (!silent) errEl.textContent = 'กรุณาใส่รหัสผ่าน'; return; }}
              try {{
                const res = await fetch('/api/admin/verify?code=' + encodeURIComponent(code));
                if (!res.ok) {{
                  errEl.textContent = res.status === 401 ? 'รหัสผ่านไม่ถูกต้อง' : 'เข้าสู่ระบบไม่สำเร็จ';
                  return;
                }}
                adminCode = code;
                localStorage.setItem(storageKey, code);
                document.getElementById('loginCard').hidden = true;
                document.getElementById('adminContent').hidden = false;
                loadEntries();
              }} catch (e) {{ if (!silent) errEl.textContent = 'เกิดข้อผิดพลาด'; }}
            }}

            function logout() {{
              adminCode = '';
              localStorage.removeItem(storageKey);
              document.getElementById('adminContent').hidden = true;
              document.getElementById('loginCard').hidden = false;
              document.getElementById('loginCode').value = '';
            }}

            // ---------- download / summary ----------
            async function download(path, filename) {{
              const code = getCode();
              statusEl.textContent = 'กำลังดาวน์โหลด...';
              try {{
                const res = await fetch(path + '?code=' + encodeURIComponent(code));
                if (!res.ok) {{
                  if (res.status === 401) {{ statusEl.textContent = 'รหัสผ่านไม่ถูกต้อง'; logout(); return; }}
                  statusEl.textContent = 'ดาวน์โหลดไม่สำเร็จ';
                  return;
                }}
                const blob = await res.blob();
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url; a.download = filename;
                document.body.appendChild(a); a.click(); a.remove();
                URL.revokeObjectURL(url);
                statusEl.textContent = 'ดาวน์โหลดสำเร็จ: ' + filename;
              }} catch (error) {{ statusEl.textContent = 'เกิดข้อผิดพลาด'; }}
            }}

            // ---------- table render ----------
            function phraseOptionsHtml(selected) {{
              return PHRASES.map(function (p, i) {{
                return '<option value="' + i + '"' + (i === selected ? ' selected' : '') + '>' + (i + 1) + '. ' + p + '</option>';
              }}).join('');
            }}

            function updateBulkCount() {{
              const n = document.querySelectorAll('#tableWrap .rowcb:checked').length;
              document.getElementById('bulkDeleteBtn').textContent = 'ลบที่เลือก (' + n + ')';
            }}

            function renderTable(entries) {{
              const wrap = document.getElementById('tableWrap');
              wrap.innerHTML = '';
              if (!entries.length) {{ wrap.innerHTML = '<p>ไม่พบรายชื่อ</p>'; updateBulkCount(); return; }}
              const table = document.createElement('table');
              table.innerHTML = '<thead><tr><th class="cb"><input type="checkbox" id="selectAll"></th><th>#</th><th>ชื่อ</th><th>ถ้อยคำ</th><th>เวลา</th><th></th></tr></thead>';
              const tbody = document.createElement('tbody');
              entries.forEach(function (e) {{
                const tr = document.createElement('tr');

                const tdCb = document.createElement('td');
                tdCb.className = 'cb';
                const cb = document.createElement('input');
                cb.type = 'checkbox';
                cb.className = 'rowcb';
                cb.setAttribute('data-id', e.id);
                cb.addEventListener('change', updateBulkCount);
                tdCb.appendChild(cb);
                tr.appendChild(tdCb);

                const tdId = document.createElement('td');
                tdId.className = 'muted-id';
                tdId.textContent = e.id;
                tr.appendChild(tdId);

                const tdName = document.createElement('td');
                const nameInput = document.createElement('input');
                nameInput.className = 'cell-name';
                nameInput.maxLength = 120;
                nameInput.value = e.name;
                tdName.appendChild(nameInput);
                tr.appendChild(tdName);

                const tdPhrase = document.createElement('td');
                const sel = document.createElement('select');
                sel.innerHTML = phraseOptionsHtml(e.phrase_index);
                tdPhrase.appendChild(sel);
                tr.appendChild(tdPhrase);

                const tdTime = document.createElement('td');
                tdTime.textContent = fmtTime(e.at);
                tr.appendChild(tdTime);

                const tdAct = document.createElement('td');
                tdAct.className = 'actions';
                const saveBtn = document.createElement('button');
                saveBtn.textContent = 'บันทึก';
                saveBtn.addEventListener('click', function () {{ saveEntry(e.id, nameInput.value, sel.value, saveBtn); }});
                const delBtn = document.createElement('button');
                delBtn.className = 'danger';
                delBtn.textContent = 'ลบ';
                delBtn.style.marginLeft = '6px';
                delBtn.addEventListener('click', function () {{ deleteEntry(e.id, e.name); }});
                tdAct.appendChild(saveBtn);
                tdAct.appendChild(delBtn);
                tr.appendChild(tdAct);

                tbody.appendChild(tr);
              }});
              table.appendChild(tbody);
              wrap.appendChild(table);

              const selAll = document.getElementById('selectAll');
              selAll.addEventListener('change', function () {{
                document.querySelectorAll('#tableWrap .rowcb').forEach(function (c) {{ c.checked = selAll.checked; }});
                updateBulkCount();
              }});
              updateBulkCount();
            }}

            function applyFilter() {{
              const term = (document.getElementById('searchInput').value || '').trim().toLowerCase();
              let filtered = allEntries;
              if (term) {{
                filtered = allEntries.filter(function (e) {{
                  return (e.name || '').toLowerCase().indexOf(term) !== -1
                    || fmtTime(e.at).toLowerCase().indexOf(term) !== -1;
                }});
              }}
              renderTable(filtered);
              setMgr('แสดง ' + filtered.length + ' / ' + allEntries.length + ' รายชื่อ');
            }}

            // ---------- CRUD ----------
            async function loadEntries() {{
              const code = getCode();
              if (!code) {{ logout(); return; }}
              setMgr('กำลังโหลด...');
              try {{
                const res = await fetch('/api/admin/entries?code=' + encodeURIComponent(code));
                if (!res.ok) {{ if (res.status === 401) {{ logout(); }} else setMgr('โหลดไม่สำเร็จ'); return; }}
                allEntries = await res.json();
                applyFilter();
              }} catch (e) {{ setMgr('เกิดข้อผิดพลาด'); }}
            }}

            async function addEntry() {{
              const code = getCode();
              const name = document.getElementById('newName').value.trim();
              const idx = parseInt(document.getElementById('newPhrase').value, 10);
              if (!name) {{ setMgr('กรุณาใส่ชื่อ'); return; }}
              setMgr('กำลังเพิ่ม...');
              try {{
                const res = await fetch('/api/admin/signatures?code=' + encodeURIComponent(code), {{
                  method: 'POST',
                  headers: {{ 'Content-Type': 'application/json' }},
                  body: JSON.stringify({{ name: name, phrase_index: idx }})
                }});
                if (!res.ok) {{ if (res.status === 401) {{ logout(); }} else setMgr('เพิ่มไม่สำเร็จ'); return; }}
                document.getElementById('newName').value = '';
                loadEntries();
              }} catch (e) {{ setMgr('เกิดข้อผิดพลาด'); }}
            }}

            async function saveEntry(id, name, phraseIdx, btn) {{
              const code = getCode();
              name = (name || '').trim();
              if (!name) {{ setMgr('ชื่อห้ามว่าง'); return; }}
              btn.disabled = true;
              try {{
                const res = await fetch('/api/admin/signatures/' + id + '?code=' + encodeURIComponent(code), {{
                  method: 'PATCH',
                  headers: {{ 'Content-Type': 'application/json' }},
                  body: JSON.stringify({{ name: name, phrase_index: parseInt(phraseIdx, 10) }})
                }});
                if (!res.ok) {{ if (res.status === 401) {{ logout(); }} else setMgr('บันทึกไม่สำเร็จ'); return; }}
                setMgr('บันทึกรายการ #' + id + ' แล้ว');
              }} catch (e) {{ setMgr('เกิดข้อผิดพลาด'); }}
              finally {{ btn.disabled = false; }}
            }}

            async function deleteEntry(id, name) {{
              if (!confirm('ลบรายการของ "' + name + '" ?')) return;
              const code = getCode();
              try {{
                const res = await fetch('/api/admin/signatures/' + id + '?code=' + encodeURIComponent(code), {{ method: 'DELETE' }});
                if (!res.ok) {{ if (res.status === 401) {{ logout(); }} else setMgr('ลบไม่สำเร็จ'); return; }}
                setMgr('ลบรายการ #' + id + ' แล้ว');
                loadEntries();
              }} catch (e) {{ setMgr('เกิดข้อผิดพลาด'); }}
            }}

            async function bulkDelete() {{
              const boxes = Array.prototype.slice.call(document.querySelectorAll('#tableWrap .rowcb:checked'));
              const ids = boxes.map(function (b) {{ return parseInt(b.getAttribute('data-id'), 10); }});
              if (!ids.length) {{ setMgr('ยังไม่ได้เลือกรายการ'); return; }}
              if (!confirm('ลบ ' + ids.length + ' รายการที่เลือก?')) return;
              const code = getCode();
              try {{
                const res = await fetch('/api/admin/signatures/bulk-delete?code=' + encodeURIComponent(code), {{
                  method: 'POST',
                  headers: {{ 'Content-Type': 'application/json' }},
                  body: JSON.stringify({{ ids: ids }})
                }});
                if (!res.ok) {{ if (res.status === 401) {{ logout(); }} else setMgr('ลบไม่สำเร็จ'); return; }}
                const d = await res.json().catch(function () {{ return {{}}; }});
                setMgr('ลบแล้ว ' + (d.deleted != null ? d.deleted : ids.length) + ' รายการ');
                loadEntries();
              }} catch (e) {{ setMgr('เกิดข้อผิดพลาด'); }}
            }}

            // ---------- init ----------
            document.getElementById('newPhrase').innerHTML = phraseOptionsHtml(0);
            document.getElementById('loginBtn').addEventListener('click', function () {{ doLogin(false); }});
            document.getElementById('loginCode').addEventListener('keydown', function (ev) {{ if (ev.key === 'Enter') doLogin(false); }});
            document.getElementById('logoutBtn').addEventListener('click', logout);
            document.getElementById('downloadCsv').addEventListener('click', function () {{ download('/api/admin/export.csv', '{EXPORT_BASENAME}.csv'); }});
            document.getElementById('downloadXlsx').addEventListener('click', function () {{ download('/api/admin/export.xlsx', '{EXPORT_BASENAME}.xlsx'); }});
            document.getElementById('loadSummary').addEventListener('click', async function () {{
              const code = getCode();
              statusEl.textContent = 'กำลังโหลดสรุป...';
              try {{
                const res = await fetch('/api/admin/summary?code=' + encodeURIComponent(code));
                const data = await res.json().catch(function () {{ return {{}}; }});
                if (!res.ok) {{ if (res.status === 401) {{ statusEl.textContent = 'รหัสผ่านไม่ถูกต้อง'; logout(); }} else statusEl.textContent = 'โหลดสรุปไม่สำเร็จ'; return; }}
                statusEl.textContent = JSON.stringify(data, null, 2);
              }} catch (e) {{ statusEl.textContent = 'เกิดข้อผิดพลาด'; }}
            }});
            document.getElementById('loadEntries').addEventListener('click', loadEntries);
            document.getElementById('addBtn').addEventListener('click', addEntry);
            document.getElementById('searchInput').addEventListener('input', applyFilter);
            document.getElementById('bulkDeleteBtn').addEventListener('click', bulkDelete);

            // auto-login ถ้าเคยจำรหัสไว้
            const stored = localStorage.getItem(storageKey) || '';
            if (stored) {{ document.getElementById('loginCode').value = stored; doLogin(true); }}
          </script>
        </body>
        </html>
        """
    ).strip()
    return html


@app.get("/")
def home() -> FileResponse:
    return FileResponse(ROOT / "index.html", media_type="text/html")


app.mount("/css", StaticFiles(directory=ROOT / "css"), name="css")
app.mount("/js", StaticFiles(directory=ROOT / "js"), name="js")
app.mount("/images", StaticFiles(directory=ROOT / "images"), name="images")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("api.index:app", host="0.0.0.0", port=int(os.getenv("PORT", "8000")), reload=False)
